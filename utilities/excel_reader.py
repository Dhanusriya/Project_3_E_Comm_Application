from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import workbook


class ExcelReader:
    @staticmethod
    def get_login_data(file_path, sheet_name):
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        data = []

        #Skip header row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)

        workbook.close()
        return data